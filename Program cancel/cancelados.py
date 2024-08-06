import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import simpledialog, messagebox

# Inicializando a lista de clientes
clientes = []

def adicionar_cliente(data, nome, cnpj, status):
    cliente = {'Data': data, 'Nome do Cliente': nome, 'CNPJ': cnpj, 'Status': status}
    clientes.append(cliente)

def carregar_dados_existentes(nome_arquivo):
    if os.path.exists(nome_arquivo):
        return pd.read_excel(nome_arquivo).to_dict('records')
    return []

def salvar_excel(nome_arquivo='clientes.xlsx'):
    # Carregar dados existentes
    dados_existentes = carregar_dados_existentes(nome_arquivo)
    
    # Adicionar novos dados aos dados existentes
    todos_os_dados = dados_existentes + clientes
    
    # Converter para DataFrame e salvar
    df = pd.DataFrame(todos_os_dados)
    df.to_excel(nome_arquivo, index=False)
    
    # Aplicar formatação condicional
    aplicar_formatacao_condicional(nome_arquivo)
    print(f'Dados salvos em {nome_arquivo}')

def aplicar_formatacao_condicional(nome_arquivo):
    workbook = load_workbook(nome_arquivo)
    sheet = workbook.active
    
    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        status_cell = row[3]
        if status_cell.value == 'Revertido':
            for cell in row:
                cell.fill = fill_verde
        elif status_cell.value == 'Cancelado':
            for cell in row:
                cell.fill = fill_vermelho
    
    workbook.save(nome_arquivo)

def coletar_dados():
    root = tk.Tk()
    root.withdraw()
    
    while True:
        nome = simpledialog.askstring("Input", "Digite o nome do cliente:")
        if nome is None:
            break
        
        cnpj = simpledialog.askstring("Input", "Digite o CNPJ do cliente (ex: 15.491.570/0001-07):")
        if cnpj is None:
            break
        
        status = simpledialog.askstring("Input", "Digite o status (1 para Revertido, 2 para Cancelado):")
        if status not in ['1', '2']:
            messagebox.showerror("Erro", "Opção inválida. Digite '1' para Revertido ou '2' para Cancelado.")
            continue
        
        status = 'Revertido' if status == '1' else 'Cancelado'
        
        data = datetime.now().strftime('%d/%m/%Y')
        adicionar_cliente(data, nome, cnpj, status)
        
        continuar = messagebox.askyesno("Continuar", "Deseja adicionar outro cliente?")
        if not continuar:
            break

    salvar_excel()

def mostrar_dados():
    nome_arquivo = 'clientes.xlsx'
    if not os.path.exists(nome_arquivo):
        messagebox.showinfo("Info", "Não há dados para mostrar.")
        return

    df = pd.read_excel(nome_arquivo)
    
    root = tk.Tk()
    root.title("Dados dos Clientes")
    
    # Configurar o tamanho inicial da janela
    root.geometry("800x600")  # Ajuste o tamanho conforme necessário
    
    # Criação da área de texto para mostrar os dados
    text = tk.Text(root, wrap='none', padx=5, pady=5)
    text.pack(expand=True, fill='both')
    
    # Configurar o cabeçalho
    text.insert(tk.END, f"{'Data':<12} {'Nome do Cliente':<30} {'CNPJ':<15} {'Status':<10}\n")
    text.insert(tk.END, '-'*67 + '\n')
    
    # Preencher os dados
    for _, row in df.iterrows():
        text.insert(tk.END, f"{row['Data']:<12} {row['Nome do Cliente']:<30} {row['CNPJ']:<15} {row['Status']:<10}\n")

    root.mainloop()

def main():
    root = tk.Tk()
    root.title("Gestão de Clientes")
    
    # Configurar o tamanho inicial da janela
    root.geometry("800x600")  # Ajuste o tamanho conforme necessário

    tk.Button(root, text="Adicionar Cliente", command=coletar_dados).pack(pady=10)
    tk.Button(root, text="Mostrar Dados", command=mostrar_dados).pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
