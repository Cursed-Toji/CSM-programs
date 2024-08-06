from flask import Flask, render_template, request, redirect, url_for
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

# Inicializando a lista de clientes
clientes = []

def adicionar_cliente(data, nome, cnpj, status, modulo, empresa, estado, motivo, analista):
    cliente = {
        'Data': data,
        'Nome do Cliente': nome,
        'CNPJ': cnpj,
        'Status': status,
        'Módulo do Cancelamento': modulo,
        'Nome da Empresa': empresa,
        'Estado': estado,
        'Motivo': motivo,
        'Analista': analista
    }
    clientes.append(cliente)

def carregar_dados_existentes(nome_arquivo):
    if os.path.exists(nome_arquivo):
        return pd.read_excel(nome_arquivo).to_dict('records')
    return []

def salvar_excel(nome_arquivo='clientes.xlsx'):
    # Carregar dados existentes
    dados_existentes = carregar_dados_existentes(nome_arquivo)
    
    # Adicionar novos dados aos dados existentes
    todos_os_dados = dados_existentes + clientes
    
    # Converter para DataFrame e salvar
    df = pd.DataFrame(todos_os_dados)
    df.to_excel(nome_arquivo, index=False)
    
    # Aplicar formatação condicional
    aplicar_formatacao_condicional(nome_arquivo)
    print(f'Dados salvos em {nome_arquivo}')

def aplicar_formatacao_condicional(nome_arquivo):
    workbook = load_workbook(nome_arquivo)
    sheet = workbook.active
    
    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        status_cell = row[3]  # Status está na quarta coluna
        if status_cell.value == 'Revertido':
            for cell in row:
                cell.fill = fill_verde
        elif status_cell.value == 'Cancelado':
            for cell in row:
                cell.fill = fill_vermelho
    
    workbook.save(nome_arquivo)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/add', methods=['GET', 'POST'])
def add_client():
    if request.method == 'POST':
        status = request.form['status']
        modulo = request.form['modulo']
        nome = request.form['nome']
        cnpj = request.form['cnpj']
        estado = request.form['estado']
        motivo = request.form['motivo']
        analista = request.form['analista']
        
        data = datetime.now().strftime('%d/%m/%Y')
        adicionar_cliente(data, nome, cnpj, status, modulo, nome, estado, motivo, analista)
        salvar_excel()
        
        return redirect(url_for('index'))
    
    return render_template('add_client.html')

@app.route('/show')
def show_data():
    nome_arquivo = 'clientes.xlsx'
    if not os.path.exists(nome_arquivo):
        return "Não há dados para mostrar."

    df = pd.read_excel(nome_arquivo)
    dados = df.to_dict(orient='records')
    
    return render_template('show_data.html', dados=dados)

@app.route('/filter', methods=['GET', 'POST'])
def filter_data():
    if request.method == 'POST':
        data_inicio = request.form['data_inicio']
        data_fim = request.form['data_fim']
        status = request.form['status']
        
        nome_arquivo = 'clientes.xlsx'
        if not os.path.exists(nome_arquivo):
            return "Não há dados para mostrar."

        df = pd.read_excel(nome_arquivo)
        df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y')
        filtro_data_inicio = datetime.strptime(data_inicio, '%d/%m/%Y')
        filtro_data_fim = datetime.strptime(data_fim, '%d/%m/%Y')
        df = df[(df['Data'] >= filtro_data_inicio) & (df['Data'] <= filtro_data_fim)]
        df = df[df['Status'] == status]
        
        dados = df.to_dict(orient='records')
        return render_template('show_data.html', dados=dados)
    
    return render_template('filter_data.html')

if __name__ == '__main__':
    app.run(debug=True)
