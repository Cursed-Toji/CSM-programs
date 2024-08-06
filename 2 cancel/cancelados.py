import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import simpledialog, messagebox
from PIL import Image, ImageTk
import requests
from io import BytesIO

# Inicializando a lista de clientes
clientes = []

def adicionar_cliente(data, nome, cnpj, status, modulo, empresa, estado, motivo, analista):
    cliente = {
        'Data': data,
        'Nome do Cliente': nome,
        'CNPJ': cnpj,
        'Status': status,
        'Módulo do Cancelamento': modulo,
        'Nome da Empresa': empresa,
        'Estado': estado,
        'Motivo': motivo,
        'Analista': analista
    }
    clientes.append(cliente)

def carregar_dados_existentes(nome_arquivo):
    if os.path.exists(nome_arquivo):
        return pd.read_excel(nome_arquivo).to_dict('records')
    return []

def salvar_excel(nome_arquivo='clientes.xlsx'):
    # Carregar dados existentes
    dados_existentes = carregar_dados_existentes(nome_arquivo)
    
    # Adicionar novos dados aos dados existentes
    todos_os_dados = dados_existentes + clientes
    
    # Converter para DataFrame e salvar
    df = pd.DataFrame(todos_os_dados)
    df.to_excel(nome_arquivo, index=False)
    
    # Aplicar formatação condicional
    aplicar_formatacao_condicional(nome_arquivo)
    print(f'Dados salvos em {nome_arquivo}')

def aplicar_formatacao_condicional(nome_arquivo):
    workbook = load_workbook(nome_arquivo)
    sheet = workbook.active
    
    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        status_cell = row[3]  # Status está na quarta coluna
        if status_cell.value == 'Revertido':
            for cell in row:
                cell.fill = fill_verde
        elif status_cell.value == 'Cancelado':
            for cell in row:
                cell.fill = fill_vermelho
    
    workbook.save(nome_arquivo)

def coletar_dados():
    root = tk.Tk()
    root.withdraw()
    
    while True:
        status = simpledialog.askstring("Input", "Digite o status (1 para Revertido, 2 para Cancelado):")
        if status not in ['1', '2']:
            messagebox.showerror("Erro", "Opção inválida. Digite '1' para Revertido ou '2' para Cancelado.")
            continue
        status = 'Revertido' if status == '1' else 'Cancelado'
        
        modulo = simpledialog.askstring("Input", "Digite o módulo do cancelamento (1 para Komunic, 2 para Acessórias):")
        if modulo not in ['1', '2']:
            messagebox.showerror("Erro", "Opção inválida. Digite '1' para Komunic ou '2' para Acessórias.")
            continue
        modulo = 'Komunic' if modulo == '1' else 'Acessórias'
        
        nome = simpledialog.askstring("Input", "Digite o nome da empresa:")
        cnpj = simpledialog.askstring("Input", "Digite o CNPJ da empresa (ex: 15.491.570/0001-07):")
        estado = simpledialog.askstring("Input", "Digite o estado:")
        motivo = simpledialog.askstring("Input", "Digite o motivo:")
        analista = simpledialog.askstring("Input", "Digite o nome do analista:")
        
        data = datetime.now().strftime('%d/%m/%Y')
        adicionar_cliente(data, nome, cnpj, status, modulo, nome, estado, motivo, analista)
        
        continuar = messagebox.askyesno("Continuar", "Deseja adicionar outro cliente?")
        if not continuar:
            break

    salvar_excel()

def mostrar_dados():
    nome_arquivo = 'clientes.xlsx'
    if not os.path.exists(nome_arquivo):
        messagebox.showinfo("Info", "Não há dados para mostrar.")
        return

    data_inicio = simpledialog.askstring("Input", "Digite a data de início (dd/mm/aaaa):")
    data_fim = simpledialog.askstring("Input", "Digite a data de fim (dd/mm/aaaa):")

    try:
        data_inicio = datetime.strptime(data_inicio, '%d/%m/%Y')
        data_fim = datetime.strptime(data_fim, '%d/%m/%Y')
    except ValueError:
        messagebox.showerror("Erro", "Data inválida. Use o formato dd/mm/aaaa.")
        return

    df = pd.read_excel(nome_arquivo)
    
    # Filtrar dados entre as datas especificadas
    df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y')
    df = df[(df['Data'] >= data_inicio) & (df['Data'] <= data_fim)]
    
    root = tk.Tk()
    root.title("Dados dos Clientes")
    
    # Configurar o tamanho inicial da janela
    root.geometry("800x600")  # Ajuste o tamanho conforme necessário
    
    # Criação da área de texto para mostrar os dados
    text = tk.Text(root, wrap='none', padx=5, pady=5)
    text.pack(expand=True, fill='both')
    
    # Configurar o cabeçalho
    text.insert(tk.END, f"{'Data':<12} {'Nome do Cliente':<20} {'CNPJ':<15} {'Status':<10} {'Módulo':<20} {'Empresa':<20} {'Estado':<10} {'Motivo':<20} {'Analista':<20}\n")
    text.insert(tk.END, '-'*135 + '\n')
    
    # Preencher os dados
    for _, row in df.iterrows():
        text.insert(tk.END, f"{row['Data'].strftime('%d/%m/%Y'):<12} {row['Nome do Cliente']:<20} {row['CNPJ']:<15} {row['Status']:<10} {row['Módulo do Cancelamento']:<20} {row['Nome da Empresa']:<20} {row['Estado']:<10} {row['Motivo']:<20} {row['Analista']:<20}\n")

    root.mainloop()

def main():
    root = tk.Tk()
    root.title("Gestão de Clientes")
    
    # Configurar o tamanho inicial da janela
    root.geometry("650x450")  # Ajuste o tamanho conforme necessário

    # Baixar a imagem da internet
    image_url = "https://i.postimg.cc/rsXsqk0Q/13339383.png"
    response = requests.get(image_url)
    img_data = response.content
    img = Image.open(BytesIO(img_data))
    img = img.resize((100, 100), Image.LANCZOS)  # Redimensionar a imagem conforme necessário
    logo = ImageTk.PhotoImage(img)
    
    # Adicionar a logo
    logo_label = tk.Label(root, image=logo)
    logo_label.image = logo  # Manter uma referência da imagem
    logo_label.pack(pady=10)

    # Adicionar o nome do sistema
    system_name_label = tk.Label(root, text="Churn Control", font=("Helvetica", 16))
    system_name_label.pack(pady=5)

    tk.Button(root, text="Adicionar Cliente", command=coletar_dados).pack(pady=10)
    tk.Button(root, text="Mostrar Dados", command=mostrar_dados).pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
